
import tkinter as tk
import tkinter.messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
 
#define the window 
root = tk.Tk()
#window dimensions
root.geometry('1080x600')

tk.Label(root, text = 'Welcome to your ExpenseTracker').grid(row=0)

#getting the user's budget
tk.Label(root, text="Enter your budget for the month").grid(row=0)
e1value = tk.IntVar()
e1 = tk.Entry(root, textvariable=e1value)
e1.grid(row=0, column=1)
seter = tk.Button(root, text='Set', fg='black', bg='lightblue')
seter.grid(row=2, column=1)
tk.Label(root, text = 'Start Recording your expenses').grid(row=5)

#lists used to store user inputs 
listi=[]
date_list = []
descrip_list  = []
#used to calculate how many times budget was exceeded
exceeded = []

#function to get the user input , row_num is used to set the placement of buttons and widgets
def expenses(row_num):

    global amountvalue
    global datevalue
    global descripvalue

    row_num +=1 

    tk.Label(root, text = 'Date').grid(row=row_num)
    datevalue= tk.StringVar()
    date = tk.Entry(root, textvariable=datevalue)
    date.grid(row=row_num, column = 2)

    row_num = row_num + 1

    tk.Label(root, text = 'Purchase Amount').grid(row=row_num)
    amountvalue = tk.IntVar()
    amount = tk.Entry(root, textvariable=amountvalue)
    amount.grid(row=row_num, column = 2)

    row_num = row_num + 1

    tk.Label(root, text='Description').grid(row=row_num)
    descripvalue = tk.StringVar()
    descrip = tk.Entry(root, textvariable=descripvalue)
    descrip.grid(row=row_num, column=2)

    row_num = row_num + 1
    
    #once values have been submitted, the submit button is commanded to call the get_values() function
    submit = tk.Button(root, text='Submit', fg='black', bg='orange', command=get_values)
    submit.grid(row=row_num, column=2)

    row_num = row_num + 1

    #button to add expenses
    add_expense = tk.Button(root, text='Add Expense', fg='black', bg='white',
                    command=lambda:expenses(row_num))
    add_expense.grid(row=row_num, column=2)

    #button to calculate total expenses and print the total to the screen
    total_expense = tk.Button(root, text = 'Calculate Total Expenses', fg = 'black', bg='white', 
                    command=lambda:print_total(True))
    total_expense.grid(row=row_num, column=3)

    row_num += 1

    #button to transfer all user data into an Excel file 
    exel = tk.Button(root, text = 'Transfer Info to Excel', fg = 'white', bg='purple', 
                    command=write_to_excel)
    exel.grid(row=row_num, column=3)

#function to get user inputted values into integers
def get_values():

    global initial_date
    global description
    global budget

    total = amountvalue.get()
    initial_date = datevalue.get()
    description = descripvalue.get()
    budget = e1value.get()
    
    total = int(total)
    budget = int(budget)

    listi.append(total)
    date_list.append(initial_date)
    descrip_list.append(description)

    print_total(False)

    return listi

#function to notify user if they meet or exceed the budget, or send a notif of total expenses

def print_total(show):

    calc = 0 

    for i in listi:
        calc = calc + i

        if show is False:
            if calc > budget:
                tk.messagebox.showinfo('Notification', 'You have exceeded your budget')
                exceeded.append('*')
            elif calc == budget:
                tk.messagebox.showinfo('Notification', 'You have met your budget')

    if show is True:
        tk.messagebox.showinfo('Total', 'Your total expenses from {} to {} are ${}.00'.format(date_list[0],date_list[-1], calc))
    

#function to write data to an excel file 
def write_to_excel():

    #create a workbook instance
    wb = Workbook()
    #load the workbook 
    wb = load_workbook('C:\\Users\\User\\Documents\\data\\myexpenses.xlsx')
    #create active worksheet 
    ws = wb.active

    #writes dates, descriptions, purchase amounts and amount of times budget exceeded to Excel spreadsheet
    placement = 'A2'
    count = 2

    for i in date_list:
        ws[placement] = i
        count = count + 1 
        placement = 'A' + str(count)
    
    count = 2 
    placement = 'B2'

    for i in descrip_list:
        ws[placement] = i
        count = count + 1 
        placement = 'B' + str(count)

    count = 2 
    placement = 'C2'
    
    for i in listi:
        ws[placement] = i 
        count = count + 1
        placement = 'C' + str(count)

    count = 2
    placement = 'D' + str(count)
    
    ws[placement] = len(exceeded)

    wb.save('C:\\Users\\User\\Documents\\data\\myexpenses.xlsx')

    tk.messagebox.showinfo('Notification', 'Transferred!')


expenses(5)
root.mainloop()